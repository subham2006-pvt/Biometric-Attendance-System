import os
import ast
import io
import cv2
import threading
import numpy as np
import mysql.connector
from datetime import datetime
from flask import Flask, render_template, Response, request, redirect, url_for, jsonify, send_file
import face_recognition
import mediapipe as mp

# Excel Styling Engines
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# --- Multi-Threaded Camera Engine Wrapper ---
class VideoCaptureThread:
    """Continuously captures raw camera buffers on a separate background thread to eliminate lagging."""
    def __init__(self, src=0):
        self.cap = cv2.VideoCapture(src)
        self.ret, self.frame = self.cap.read()
        self.running = True
        self.thread = threading.Thread(target=self.update, args=(), daemon=True)
        self.thread.start()

    def update(self):
        while self.running:
            ret, frame = self.cap.read()
            if ret:
                self.ret, self.frame = ret, frame

    def read(self):
        return self.ret, self.frame.copy() if self.frame is not None else None

    def stop(self):
        self.running = False
        self.cap.release()

# Spin up our multi-threaded camera instance
camera = VideoCaptureThread(0)
PALM_TRIGGER_FLAG = False

# --- MySQL Database Configuration ---
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",  
        database="attendance_system"
    )

# --- MediaPipe Configuration Setup ---
mp_hands = mp.solutions.hands
hands = mp_hands.Hands(static_image_mode=False, max_num_hands=1, min_detection_confidence=0.6)

def load_known_faces():
    known_encodings, known_names, known_rolls = [], [], []
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT roll_no, name, encoding FROM students")
        for row in cursor.fetchall():
            roll_no, name, enc_str = row
            known_encodings.append(np.array(ast.literal_eval(enc_str)))
            known_names.append(name)
            known_rolls.append(roll_no)
        cursor.close(); conn.close()
    except Exception as e:
        print(f"Database sync error: {e}")
    return known_encodings, known_names, known_rolls

def process_attendance_transaction(roll_no, name):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')
        now_time = datetime.now()
        now_str = now_time.strftime('%H:%M:%S')

        cursor.execute(
            "SELECT id, in_time, out_time FROM attendance_logs WHERE roll_no = %s AND log_date = %s ORDER BY id DESC LIMIT 1",
            (roll_no, today)
        )
        last_log = cursor.fetchone()

        if not last_log or last_log[2] != "Pending":
            sql = "INSERT INTO attendance_logs (log_date, roll_no, name, in_time, out_time) VALUES (%s, %s, %s, %s, 'Pending')"
            cursor.execute(sql, (today, roll_no, name, now_str))
            conn.commit()
            cursor.close(); conn.close()
            return f"Logged In: {name}", (0, 255, 0)

        log_id, in_time_str, _ = last_log
        in_time_parsed = datetime.strptime(f"{today} {in_time_str}", '%Y-%m-%d %H:%M:%S')
        elapsed_hours = (now_time - in_time_parsed).total_seconds() / 3600.0

        if elapsed_hours < 4.0:
            cursor.close(); conn.close()
            return "Lockout: < 4 Hrs!", (0, 0, 255)
        else:
            cursor.execute("UPDATE attendance_logs SET out_time = %s WHERE id = %s", (now_str, log_id))
            conn.commit()
            cursor.close(); conn.close()
            return f"Logged Out: {name}", (255, 165, 0)
    except Exception as e:
        print(f"Transaction failure: {e}")
        return "System DB Error", (0, 0, 255)

def analyze_hand_gestures(frame):
    global PALM_TRIGGER_FLAG
    # Downscale for faster gesture detection performance
    small_hand_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
    rgb_frame = cv2.cvtColor(small_hand_frame, cv2.COLOR_BGR2RGB)
    results = hands.process(rgb_frame)
    if results.multi_hand_landmarks:
        for hand_landmarks in results.multi_hand_landmarks:
            lms = hand_landmarks.landmark
            
            # ✋ Open Palm Analysis
            palm_check = (lms[4].y < lms[3].y) and (lms[8].y < lms[6].y) and \
                         (lms[12].y < lms[10].y) and (lms[16].y < lms[14].y) and \
                         (lms[20].y < lms[18].y)
            if palm_check: return "PALM"
            
            # 👍 Thumbs Up Analysis
            thumb_check = lms[4].y < lms[3].y < lms[2].y
            fingers_curled = (lms[8].x > lms[6].x if lms[5].x < lms[17].x else lms[8].x < lms[6].x) and \
                             (lms[12].y > lms[10].y) and (lms[16].y > lms[14].y) and (lms[20].y > lms[18].y)
            if thumb_check and fingers_curled: return "THUMBS_UP"
    return None

def generate_video_stream():
    global PALM_TRIGGER_FLAG
    frame_count = 0
    hud_message = ""
    hud_color = (255, 255, 255)
    message_timeout = 0
    
    # Store persistent overlays between evaluation ticks to stop lagging freezes
    cached_face_boxes = []
    
    while True:
        success, frame = camera.read()
        if not success or frame is None: 
            continue
            
        frame_count += 1
        if message_timeout > 0:
            message_timeout -= 1
            if message_timeout == 0: hud_message = ""

        # ONLY execute biometric arrays every 4th frame (Saves huge CPU cycles)
        if frame_count % 4 == 0:
            # Drop size resolution down to 25% for lightning-fast matching speeds
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
            known_encodings, known_names, known_rolls = load_known_faces()
            
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            
            cached_face_boxes = []
            identity = "Unknown Profile"
            roll_number = None

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
                if True in matches:
                    first_match_index = matches.index(True)
                    identity = known_names[first_match_index]
                    roll_number = known_rolls[first_match_index]
                    
                # Store coordinates upscaled back to 100% video display layer resolution
                cached_face_boxes.append((top*4, right*4, bottom*4, left*4, identity, roll_number))

            gesture = analyze_hand_gestures(frame)
            if gesture == "PALM" and identity == "Unknown Profile":
                PALM_TRIGGER_FLAG = True
            elif gesture == "THUMBS_UP" and identity != "Unknown Profile":
                hud_message, hud_color = process_attendance_transaction(roll_number, identity)
                message_timeout = 30

        # Render tracking bounding boxes smoothly out of our saved cache
        for (top, right, bottom, left, identity, roll_number) in cached_face_boxes:
            box_color = (0, 255, 0) if identity != "Unknown Profile" else (0, 0, 255)
            cv2.rectangle(frame, (left, top), (right, bottom), box_color, 2)
            cv2.putText(frame, identity, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 1)

        if hud_message:
            cv2.putText(frame, hud_message, (30, 60), cv2.FONT_HERSHEY_SIMPLEX, 1.1, hud_color, 3)

        ret, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')

# --- HTTP Web App System Endpoints ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_video_stream(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/check_palm')
def check_palm():
    global PALM_TRIGGER_FLAG
    return jsonify({"palm_detected": PALM_TRIGGER_FLAG})

@app.route('/reset_palm')
def reset_palm():
    global PALM_TRIGGER_FLAG
    PALM_TRIGGER_FLAG = False
    return jsonify({"status": "reset"})

@app.route('/api/students', methods=['GET'])
def api_get_students():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT roll_no, name FROM students ORDER BY name ASC")
        students = cursor.fetchall()
        cursor.execute("SELECT log_date, roll_no, name, in_time, out_time FROM attendance_logs ORDER BY id DESC")
        logs = cursor.fetchall()
        cursor.close(); conn.close()
        return jsonify({"students": students, "logs": logs})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/register', methods=['POST'])
def register():
    global PALM_TRIGGER_FLAG
    roll_no = request.form.get('roll_no')
    name = request.form.get('name')
    
    encodings = []
    for _ in range(15):  # Enhanced range validation sweep 
        success, frame = camera.read()
        if not success or frame is None: continue
        encodings = face_recognition.face_encodings(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        if len(encodings) > 0: break

    if len(encodings) == 0:
        return jsonify({"status": "error", "message": "Stable face vectors not detected. Try again!"})
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO students (roll_no, name, encoding) VALUES (%s, %s, %s)", (roll_no, name, str(list(encodings[0]))))
        conn.commit(); cursor.close(); conn.close()
        PALM_TRIGGER_FLAG = False
        return redirect(url_for('index'))
    except mysql.connector.Error as err:
        return jsonify({"status": "error", "message": f"Database allocation crash: {err}"})

@app.route('/export/excel', methods=['GET'])
def export_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT log_date, roll_no, name, in_time, out_time FROM attendance_logs ORDER BY log_date DESC, in_time DESC")
        logs = cursor.fetchall()
        cursor.close(); conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Attendance Ledger"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") 
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        accent_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid") 

        font_title = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10, color="334155")
        font_bold_body = Font(name="Segoe UI", size=10, bold=True, color="047857")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin_border_side = Side(border_style="thin", color="E2E8F0")
        border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        ws['A1'] = "BIOMETRIC ATTENDANCE SYSTEM - SYSTEM LEDGER REPORT"
        ws['A1'].font = font_title
        ws.row_dimensions[1].height = 35

        headers = ["Log Date", "Roll Number", "Full Name", "Check-In Time", "Check-Out Time", "Shift Status"]
        ws.append([]) 
        ws.append(headers)
        ws.row_dimensions[3].height = 26

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_cell

        current_row = 4
        for log in logs:
            log_date, roll_no, name, in_time, out_time = log
            status_text = "Completed" if out_time != "Pending" else "Active Shift"
            row_data = [str(log_date), roll_no, name, in_time, out_time, status_text]
            
            ws.append(row_data)
            ws.row_dimensions[current_row].height = 20
            
            row_fill = zebra_fill if current_row % 2 == 0 else white_fill
            if status_text == "Active Shift": row_fill = accent_fill

            for col_idx in range(1, 7):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = row_fill
                cell.font = font_body if col_idx != 6 else font_bold_body
                cell.border = border_cell
                cell.alignment = align_center if col_idx in [1, 2, 4, 5, 6] else align_left
            current_row += 1

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1: continue 
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = f"Biometric_Attendance_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(file_stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"status": "error", "message": f"Excel runtime error: {e}"}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)